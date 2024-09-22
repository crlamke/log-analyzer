"""
Excel Utilities - This module uses openpyxl to provide functionality to read and write
Microsoft Excel files.
"""

"""
@author: Chris Lamke
"""

import os
from openpyxl import Workbook # Using https://openpyxl.readthedocs.io/en/stable/tutorial.html for write to Excel
from openpyxl.styles import *

""" XLSFormat class that contains formatting details for Excel data """
class XLSFormat:
    wb = None
    is_bold = False
    is_italic = False
    row_height = 0
    col_width = 0
    font = None

    def __init__(self, excel_file_dir, excel_file_name):
        self.font = Font(bold=True)

    def get_bold(self):
        pass



""" XLSDoc class that creates, reads, and writes Excel files """
class XLSDoc:
    wb = None
    excel_file_name = ""
    excel_file_dir = ""
    excel_file = ""

    def __init__(self, excel_file_dir, excel_file_name):
        self.excel_file_name = excel_file_name
        self.excel_file_dir = excel_file_dir
        self.excel_file = os.path.join(excel_file_dir, excel_file_name)

        self.wb = Workbook()
        ws = self.wb.active

    def save_doc(self):
        self.wb.save(self.excel_file)

    def create_worksheet(self, sheet_name, index):
        self.wb.create_sheet(sheet_name, index) 
        return self.wb[sheet_name]

    def delete_worksheet(self, sheet_name):
        self.wb.remove(self.wb[sheet_name])

    def get_worksheet_by_name(self, sheet_name):
        return self.wb[sheet_name]

    def write_cell(self, ws, ws_row, ws_col, cell_value):
        ws.cell(row=ws_row, column=ws_col).value = cell_value 

    def write_cell_with_format(self, ws, ws_row, ws_col, cell_value, font):
        ws.cell(row=ws_row, column=ws_col).value = cell_value 
        ws.cell(row=ws_row, column=ws_col).font = Font(bold=True)
        ws.cell(row=ws_row, column=ws_col).alignment = Alignment(
            horizontal="center", vertical="center")
